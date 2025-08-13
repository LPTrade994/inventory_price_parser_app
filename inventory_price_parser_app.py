# inventory_price_parser_app.py
# ---------------------------------------------------------
# Streamlit web-app per caricare un file inventario (ITALIA.xlsx
# o inventario.txt) e un file prezzi d'acquisto (PREZZI ACQUISTO.xlsx
# o acquisti.txt) e abbinarli
# tramite SKU. Il prezzo di riferimento usato è "Prezzo medio".
# Esecuzione locale:  
#   streamlit run inventory_price_parser_app.py
# ---------------------------------------------------------

import io
import re

import openpyxl
import pandas as pd
import numpy as np
import streamlit as st

CATEGORY_MAP = {
    "Videogiochi - Giochi e accessori": {"referral": 15.0, "closing": 0.81},
    "Videogiochi - Console": {"referral": 8.0, "closing": 0.81},
    "Libri": {"referral": 15.0, "closing": 1.01},
    "Musica": {"referral": 15.0, "closing": 0.81},
    "Video e DVD": {"referral": 15.0, "closing": 0.81},
    "Software": {"referral": 15.0, "closing": 0.81},
    "_default": {"referral": 15.0, "closing": 0.00},
}
DST_PCT = 3.0  # Italia

COUNTRY_CODES = {
    "AE",
    "AU",
    "BE",
    "BR",
    "CA",
    "CN",
    "DE",
    "ES",
    "FR",
    "GB",
    "IT",
    "JP",
    "MX",
    "NL",
    "PL",
    "SE",
    "TR",
    "UK",
    "US",
    "IE",
}


# ---------------------------------------------------------
# Sidebar – caricamento file e opzioni
# ---------------------------------------------------------
st.set_page_config(page_title="Inventory & Price Parser", layout="wide")
st.title("📦 Inventory & Price Parser")

with st.sidebar:
    st.header("⚙️ Opzioni")
    parse_option = st.checkbox(
        "Ignora suffisso dopo l'ultimo '-' nello SKU (es. 4556415-2-XY → 4556415-2)",
        value=True,
    )

    inv_file = st.file_uploader(
        "📥 File inventario (es. ITALIA.xlsx o inventario.txt)",
        type=["xlsx", "xls", "txt"],
        key="inv_uploader",
    )

    price_file = st.file_uploader(
        "📥 File acquisti (es. PREZZI ACQUISTO.xlsx o acquisti.txt)",
        type=["xlsx", "xls", "txt"],
        key="price_uploader",
    )

    export_file = st.file_uploader(
        "📥 Flat-file Amazon (export) • Non Automated SKUs",
        type=["xlsx", "xls", "csv", "tsv"],
        key="export_uploader",
    )

    download_name = st.text_input(
        "Nome file Excel da scaricare", value="inventario_con_prezzi.xlsx"
    )

# ---------------------------------------------------------
# Funzioni di utilità
# ---------------------------------------------------------
@st.cache_data(show_spinner=False)
def load_excel(uploaded_file):
    if uploaded_file is None:
        return None

    name = uploaded_file.name.lower()
    if name.endswith(".txt"):
        uploaded_file.seek(0)
        content = uploaded_file.read().decode("latin-1", errors="ignore")
        # prova TSV e CSV
        try:
            df = pd.read_csv(io.StringIO(content), sep="\t", decimal=",")
        except Exception:
            df = pd.read_csv(io.StringIO(content), sep=";", decimal=",")
        return df

    # Excel
    try:
        df = pd.read_excel(uploaded_file, engine="openpyxl")
    except Exception:
        uploaded_file.seek(0)
        df = pd.read_excel(uploaded_file)
    return df


def _to_kebab(s: str) -> str:
    s = "" if s is None else str(s)
    s = s.strip().lower()
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"[^a-z0-9\-]", "", s)
    return s


def load_amazon_template(uploaded_file):
    """Carica un file *template* Automate Pricing (doppio header fisso)."""
    uploaded_file.seek(0)
    wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet_name = next((n for n in [
        "Modello assegnazione prezzo",
        "Modello di assegnazione prezzo",
        "Modello di assegnazione del prezzo",
        "Esempio",
    ] if n in wb.sheetnames), wb.sheetnames[0])

    ws = wb[sheet_name]
    data = list(ws.values)
    if len(data) < 2:
        return pd.DataFrame()

    # riga 2 (index 1) = nomi macchina nella maggior parte dei template
    header_raw = data[1]
    columns = [_to_kebab(h) for h in header_raw]
    rows = [list(row) for row in data[2:]]
    df = pd.DataFrame(rows, columns=columns)

    df = df.dropna(how="all")
    if "sku" in df.columns:
        df["sku"] = df["sku"].astype(str).str.strip()
        df = df[df["sku"] != ""]

    # normalizza stringhe
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].astype(str).str.strip()

    # alias e rinomina
    rename_map = {"sku": "SKU", "seller-sku": "SKU", "current-selling-price": "Prezzo"}
    df = df.rename(columns=rename_map)

    # coercizioni numeriche
    price_cols = [c for c in df.columns if re.search(r"price|points", c)]
    for col in price_cols:
        df[col] = pd.to_numeric(
            pd.Series(df[col]).astype(str).str.replace(",", "."), errors="coerce"
        )
        df.loc[df[col] < 0, col] = np.nan

    if "sales-rank" in df.columns:
        df["sales-rank"] = pd.to_numeric(
            pd.Series(df["sales-rank"]).astype(str).str.replace(",", "."), errors="coerce"
        ).astype("Int64")

    if "rule-action" in df.columns:
        df["rule-action"] = df["rule-action"].astype(str).str.strip().str.upper()
        valid_actions = {"START", "STOP"}
        bad = df.loc[~df["rule-action"].isin(valid_actions) & df["rule-action"].notna(), "rule-action"].unique()
        if len(bad) > 0:
            raise ValueError(f"Valori 'rule-action' non validi: {', '.join(bad)}")

    if "country-code" in df.columns:
        df["country-code"] = df["country-code"].astype(str).str.strip().str.upper()
        bad_cc = df.loc[~df["country-code"].isin(COUNTRY_CODES) & df["country-code"].notna(), "country-code"].unique()
        if len(bad_cc) > 0:
            raise ValueError(f"Codici paese non validi: {', '.join(bad_cc)}")

    return df.reset_index(drop=True)


def load_amazon_export(uploaded_file):
    """Parser *robusto* del flat-file **export** (header variabile, CSV/XLSX/TSV).

    - Scansiona le prime ~30 righe per trovare la riga che contiene 'sku' (o 'seller-sku').
    - Usa quella riga come intestazione effettiva (in kebab-case) e legge i record successivi.
    - Supporta Excel (qualsiasi nome foglio) e CSV/TSV (senza header).
    """
    name = uploaded_file.name.lower()

    def _build_df_from_matrix(matrix):
        # trova riga header
        header_idx = None
        header_row = None
        for i in range(min(30, len(matrix))):
            row_vals = ["" if v is None else str(v) for v in matrix[i]]
            keb = [_to_kebab(v) for v in row_vals]
            if "sku" in keb or "seller-sku" in keb:
                header_idx = i
                header_row = keb
                break
        if header_idx is None:
            # fallback: prova seconda riga come nei template
            header_idx = 1 if len(matrix) > 1 else 0
            header_row = [_to_kebab(v) for v in (matrix[header_idx] if matrix else [])]

        data_rows = [list(r) for r in matrix[header_idx + 1 :]]
        df = pd.DataFrame(data_rows, columns=header_row)
        # pulizie
        for col in df.select_dtypes(include="object").columns:
            df[col] = df[col].astype(str).str.strip()
        # alias sku
        ren = {"sku": "SKU", "seller-sku": "SKU", "current-selling-price": "Prezzo"}
        df = df.rename(columns=ren)
        return df

    if name.endswith((".xlsx", ".xls")):
        uploaded_file.seek(0)
        wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        # scegli foglio preferito se presente, altrimenti il primo
        sheet_name = next((n for n in [
            "Modello assegnazione prezzo",
            "Modello di assegnazione prezzo",
            "Modello di assegnazione del prezzo",
            "Esempio",
        ] if n in wb.sheetnames), wb.sheetnames[0])
        ws = wb[sheet_name]
        matrix = list(ws.values)
        df = _build_df_from_matrix(matrix)
    else:
        # CSV/TSV senza header deterministico
        uploaded_file.seek(0)
        content = uploaded_file.read().decode("utf-8", errors="ignore")
        # tenta separatori comuni
        frames = []
        for sep in [",", "\t", ";", "|"]:
            try:
                tmp = pd.read_csv(io.StringIO(content), sep=sep, header=None, dtype=str)
                frames.append(tmp)
            except Exception:
                pass
        if not frames:
            raise ValueError("Impossibile leggere il flat-file export (CSV/TSV)")
        # scegli quello con più colonne (euristica)
        df_raw = max(frames, key=lambda d: d.shape[1])
        matrix = df_raw.fillna("").values.tolist()
        df = _build_df_from_matrix(matrix)

    # coercizioni numeriche standard
    num_cols = [
        "minimum-seller-allowed-price",
        "maximum-seller-allowed-price",
        "current-selling-price",
        "buybox-landed-price",
        "lowest-landed-price",
        "sales-rank",
    ]
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(pd.Series(df[c]).astype(str).str.replace(",", "."), errors="coerce")

    # normalizza campi chiave
    if "rule-action" in df.columns:
        df["rule-action"] = df["rule-action"].astype(str).str.upper().replace({"": "START", "nan": "START"})
    if "country-code" in df.columns:
        df["country-code"] = df["country-code"].astype(str).str.upper()
    if "currency-code" in df.columns:
        df["currency-code"] = df["currency-code"].astype(str).str.upper()

    # filtra righe vuote sku
    if "SKU" in df.columns:
        df = df[df["SKU"].astype(str).str.strip() != ""]

    return df.reset_index(drop=True)


def normalize_sku(series: pd.Series, parse_suffix: bool) -> pd.Series:
    """Restituisce SKU come stringa.

    Se ``parse_suffix`` è ``True`` rimuove solo il suffisso dopo l'ultimo
    trattino (``-``), ad esempio ``ABC-1-2`` diventa ``ABC-1``.
    """
    series = series.astype(str).str.strip()
    if parse_suffix:
        return series.str.rsplit("-", n=1).str[0]
    return series


def get_merged_inventory(
    inventory_df: pd.DataFrame,
    purchase_df: pd.DataFrame,
    parse_suffix: bool,
) -> tuple[pd.DataFrame, str]:
    """Esegue il merge mantenendo la colonna Categoria se presente.

    Restituisce anche il nome della colonna usata come SKU originale
    nell'inventario.
    """
    inv_key_candidates = [
        c for c in inventory_df.columns if c.upper() in {"SKU", "CODICE(ASIN)", "CODICE", "ASIN"}
    ]
    price_key_candidates = [
        c for c in purchase_df.columns if c.upper() in {"CODICE", "SKU", "CODICE(ASIN)"}
    ]
    if not inv_key_candidates or not price_key_candidates:
        st.error(
            "Assicurati che entrambi i file contengano almeno una colonna chiave tra: "
            "SKU, CODICE(ASIN), CODICE, ASIN."
        )
        st.stop()

    inv_key = st.selectbox("Colonna SKU nell'inventario", inv_key_candidates, index=0)
    price_key = st.selectbox("Colonna SKU nel file acquisti", price_key_candidates, index=0)

    inventory_df["_SKU_KEY_"] = normalize_sku(inventory_df[inv_key], parse_suffix)
    purchase_df["_SKU_KEY_"] = normalize_sku(purchase_df[price_key], parse_suffix)

    # trova colonna prezzo più adatta
    candidate_price_cols = [
        c
        for c in purchase_df.columns
        if "prezzo" in c.lower() and "medio" in c.lower()
        or c.lower().strip() in {"prezzo", "prezzo medio", "prezzo medio (€)"}
    ]
    if not candidate_price_cols:
        st.error(
            "Colonna del prezzo d'acquisto non trovata (es. 'Prezzo medio', 'Prezzo')."
        )
        st.stop()
    price_col = candidate_price_cols[0]

    # mantiene "Categoria" se presente
    cat_cols = [c for c in purchase_df.columns if c.lower().strip() == "categoria"]

    optional_cols = [
        c
        for c in purchase_df.columns
        if c.lower().strip()
        in {
            "quantita'",
            "quantità",
            "prezzo",
            "prezzo medio",
            "codice(asin)",
            "asin",
            "sku",
            "country-code",
            "currency-code",
            "rule-name",
            "rule-action",
            "business-rule-name",
            "business-rule-action",
        }
    ]
    subset_cols = ["_SKU_KEY_", price_col] + optional_cols
    if cat_cols:
        purchase_df = purchase_df.rename(columns={cat_cols[0]: "Categoria"})
        subset_cols.append("Categoria")

    merged = inventory_df.merge(
        purchase_df[subset_cols], on="_SKU_KEY_", how="left", suffixes=("", "_acquisto")
    )
    merged = merged.rename(columns={price_col: "Prezzo medio acquisto (€)"})
    return merged, inv_key


def calc_min_price(
    row: pd.Series,
    referral_pct: float,
    closing_fee: float,
    dst_pct: float,
    ship_cost: float,
    vat_pct: float,
    margin_pct: float,
):
    """Calcola il prezzo minimo ivato che soddisfi margine/fee/costi.

    Formula:
        P = k * (C + S + (1 + d)*F) / (1 - k*((1 + d)*R + M))

    dove:
        k = 1 + IVA
        C = costo acquisto
        S = costo spedizione
        F = fee di chiusura
        R = referral fee
        d = DST
        M = margine target
    """
    cost = pd.to_numeric(row.get("Prezzo medio acquisto (€)"), errors="coerce")
    if not np.isfinite(cost) or cost <= 0:
        return None

    r = referral_pct / 100.0
    d = dst_pct / 100.0
    m = margin_pct / 100.0
    v = vat_pct / 100.0

    ship_cost = float(ship_cost or 0.0)
    closing_fee = float(closing_fee or 0.0)

    k = 1 + v

    numerator = k * (cost + ship_cost + (1 + d) * closing_fee)
    denom     = 1 - k * ((1 + d) * r + m)

    if denom <= 0:
        return None  # parametri impossibili (fee+margine troppo alti)

    return round(numerator / denom, 2)


def build_flatfile(df: pd.DataFrame, sku_col: str) -> pd.DataFrame:
    field_names = [
        "sku",
        "minimum-seller-allowed-price",
        "maximum-seller-allowed-price",
        "country-code",
        "currency-code",
        "rule-name",
        "rule-action",
        "business-rule-name",
        "business-rule-action",
    ]

    header_desc = [
        "SKU",
        "Min price",
        "Max price",
        "Country code",
        "Currency code",
        "Rule name",
        "Rule action",
        "Business rule name",
        "Business rule action",
    ]

    data = {
        "sku": df.get(sku_col, df.get("SKU")),
        "minimum-seller-allowed-price": df.get("minimum-seller-allowed-price", df.get("Prezzo minimo suggerito (€)")),
        "maximum-seller-allowed-price": df.get("maximum-seller-allowed-price", ""),
        "country-code": df.get("country-code", "IT"),
        "currency-code": df.get("currency-code", "EUR"),
        "rule-name": df.get("rule-name", "Rule1"),
        "rule-action": df.get("rule-action", "START"),  # default in UPPERCASE
        "business-rule-name": df.get("business-rule-name", ""),
        "business-rule-action": df.get("business-rule-action", ""),
    }

    df_data = pd.DataFrame(data)
    df_full = pd.concat(
        [pd.DataFrame([header_desc], columns=field_names), pd.DataFrame([field_names], columns=field_names), df_data],
        ignore_index=True,
    )
    return df_full


def make_flatfile_bytes(df: pd.DataFrame) -> io.BytesIO:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, header=False, sheet_name="Modello assegnazione prezzo")
    output.seek(0)
    return output


def highlight_below(row):
    min_val = pd.to_numeric(row.get("Prezzo minimo suggerito (€)"), errors="coerce")
    price_val = pd.to_numeric(row.get("Prezzo"), errors="coerce")
    if np.isfinite(min_val) and np.isfinite(price_val) and min_val > price_val:
        return ["background-color: lightcoral"] * len(row)
    return [""] * len(row)


# ---------------------------------------------------------
# Caricamento dati
# ---------------------------------------------------------
inventory_df = load_excel(inv_file)
purchase_df = load_excel(price_file)

# rinomina colonne chiave se presenti
rename_map = {"sku": "SKU", "seller-sku": "SKU", "current-selling-price": "Prezzo"}
if inventory_df is not None:
    inventory_df = inventory_df.rename(columns=rename_map)
if purchase_df is not None:
    purchase_df = purchase_df.rename(columns=rename_map)

# Early-return rivisto:
# - Se NON c'è l'export Amazon: servono inventario + acquisti.
# - Se c'è l'export Amazon: puoi lavorare anche solo con **inventario** (costi dall'inventario)
#   oppure solo con **acquisti** (costi dal file acquisti). Qui attiviamo i rami dedicati.
if export_file is None and (inventory_df is None or purchase_df is None):
    st.info("👈 Carica inventario + acquisti, oppure carica flat-file Amazon + inventario (o + acquisti).")
    st.stop()

# -------------------------------------------------------------------
# SOLO EXPORT + INVENTARIO: flat-file Amazon + file inventario (senza acquisti)
# -------------------------------------------------------------------
if export_file is not None and inventory_df is not None and purchase_df is None:
    st.subheader("Non Automated SKUs – da export Amazon (solo export + inventario)")

    # Leggo l'export con header dinamico
    try:
        export_df = load_amazon_export(export_file)
    except Exception as e:
        st.error(f"Errore nel parsing del flat-file export: {e}")
        st.stop()

    if "SKU" not in export_df.columns:
        st.error("Il flat-file export non contiene 'SKU'.")
        st.stop()

    export_df["_SKU_KEY_"] = normalize_sku(export_df["SKU"], parse_option)

    # Individuo chiave SKU nell'inventario
    inv_key_candidates = [c for c in inventory_df.columns if c.upper() in {"SKU","CODICE(ASIN)","CODICE","ASIN"}]
    if not inv_key_candidates:
        st.error("Nel file inventario manca la colonna SKU/CODICE/CODICE(ASIN)/ASIN.")
        st.stop()
    inv_key_exp = st.selectbox("Colonna SKU nell'inventario", inv_key_candidates, index=0, key="expinv_invkey")
    inventory_df["_SKU_KEY_"] = normalize_sku(inventory_df[inv_key_exp], parse_option)

    # Scelta della colonna COSTO nell'inventario
    priority = [
        "Prezzo medio acquisto (€)",
        "Prezzo medio acquisto",
        "Prezzo acquisto",
        "Prezzo d'acquisto",
        "Costo acquisto",
        "Costo",
        "Costo unitario",
        "Costo medio",
        "Prezzo medio",
    ]
    # Candidati per nome e per contenuto (acquisto/costo)
    name_hits = [c for c in inventory_df.columns if any(p.lower() in c.lower() for p in ["acquisto","costo","cost","purchase"]) ]
    candidates = [c for c in priority if c in inventory_df.columns]
    for c in name_hits:
        if c not in candidates:
            candidates.append(c)
    if not candidates:
        # ultima spiaggia: tutte le colonne numeriche
        num_cols = [c for c in inventory_df.columns if pd.api.types.is_numeric_dtype(inventory_df[c])]
        if not num_cols:
            st.error("Non trovo una colonna costo nell'inventario. Aggiungi una colonna costo o carica anche il file acquisti.")
            st.stop()
        candidates = num_cols

    cost_col = st.selectbox("Colonna costo d'acquisto nell'inventario", candidates, index=0, key="expinv_costcol")

    # Categoria opzionale
    cat_col = None
    for c in inventory_df.columns:
        if c.lower().strip() == "categoria":
            cat_col = c
            break

    inv_subset_cols = ["_SKU_KEY_", cost_col] + ([cat_col] if cat_col else [])
    inv_subset = inventory_df[inv_subset_cols].copy()
    inv_subset = inv_subset.rename(columns={cost_col: "Prezzo medio acquisto (€)", cat_col: "Categoria"} if cat_col else {cost_col: "Prezzo medio acquisto (€)"})

    # Merge export ⟷ inventario (per avere il costo)
    merged_export = export_df.merge(
        inv_subset,
        on="_SKU_KEY_", how="left", suffixes=("", "_inv")
    )

    # Parametri (indipendenti dal file acquisti)
    st.markdown("### Parametri costi e margini (solo export + inventario)")
    cats = list(CATEGORY_MAP.keys())
    selected_cat_exp = st.selectbox("Categoria predefinita", cats, index=0, key="expinv_cat")
    defaults_exp = CATEGORY_MAP.get(selected_cat_exp, CATEGORY_MAP["_default"])
    referral_fee_pct_exp = st.number_input("% Commissione Amazon", value=defaults_exp["referral"], min_value=0.0, key="expinv_ref")
    shipping_cost_exp     = st.number_input("Costo spedizione €", value=0.0, min_value=0.0, key="expinv_ship")
    vat_pct_exp           = st.number_input("IVA %", value=22.0, min_value=0.0, step=0.1, key="expinv_vat")
    margin_pct_exp        = st.number_input("Margine desiderato %", value=20.0, min_value=0.0, key="expinv_margin")
    closing_fee_exp       = defaults_exp.get("closing", 0.0)

    # Calcolo minimo
    merged_export["Prezzo medio acquisto (€)"] = pd.to_numeric(merged_export["Prezzo medio acquisto (€)"], errors="coerce")
    merged_export["Prezzo minimo suggerito (€)"] = merged_export.apply(
        calc_min_price,
        axis=1,
        referral_pct=referral_fee_pct_exp,
        closing_fee=closing_fee_exp,
        dst_pct=DST_PCT,
        ship_cost=shipping_cost_exp,
        vat_pct=vat_pct_exp,
        margin_pct=margin_pct_exp,
    )

    # Opzioni flat-file
    only_missing_min_exp = st.checkbox("Mostra solo righe senza 'minimum-seller-allowed-price' nel file", value=True, key="expinv_missmin")
    overwrite_min_exp    = st.checkbox("Sovrascrivi 'minimum-seller-allowed-price' se già presente", value=False, key="expinv_overwrite")

    default_rule_base_exp = st.text_input("Rule name di default", value="AUTO", key="expinv_rulename")
    today_str_exp = pd.Timestamp.today().strftime("%Y%m%d")

    if "rule-name" not in merged_export.columns:
        merged_export["rule-name"] = ""
    if "country-code" not in merged_export.columns:
        merged_export["country-code"] = "IT"
    if "currency-code" not in merged_export.columns:
        merged_export["currency-code"] = "EUR"

    merged_export["rule-action"] = merged_export.get("rule-action", "START")
    merged_export["rule-action"] = merged_export["rule-action"].astype(str).str.upper().replace({"": "START"})

    def _mk_rule_name_exp(row):
        rn = str(row.get("rule-name") or "").strip()
        if rn:
            return rn
        cc = str(row.get("country-code") or "IT").upper()
        return f"{default_rule_base_exp}-{cc}-{today_str_exp}"

    merged_export["rule-name"] = merged_export.apply(_mk_rule_name_exp, axis=1)

    # Filtri e riempimento minimo
    if "minimum-seller-allowed-price" in merged_export.columns and only_missing_min_exp:
        view_exp = merged_export[merged_export["minimum-seller-allowed-price"].isna()].copy()
    else:
        view_exp = merged_export.copy()

    if overwrite_min_exp or "minimum-seller-allowed-price" not in view_exp.columns:
        view_exp["minimum-seller-allowed-price"] = view_exp["Prezzo minimo suggerito (€)"]
    else:
        view_exp["minimum-seller-allowed-price"] = view_exp["minimum-seller-allowed-price"].where(
            view_exp["minimum-seller-allowed-price"].notna(), view_exp["Prezzo minimo suggerito (€)"]
        )

    ff_out_exp = view_exp.dropna(subset=["SKU", "Prezzo medio acquisto (€)", "minimum-seller-allowed-price"])  # richiedo costo valido

    st.dataframe(
        ff_out_exp[[
            "SKU","country-code","currency-code","rule-name","rule-action",
            "minimum-seller-allowed-price","Prezzo medio acquisto (€)","Prezzo minimo suggerito (€)"
        ]].head(100),
        use_container_width=True, hide_index=True
    )

    st.download_button(
        "💾 Scarica Flat-File (compilato da export)",
        data=make_flatfile_bytes(build_flatfile(ff_out_exp, "SKU")),
        file_name="AutomatePricing_FromExport.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.stop()  # Evita di proseguire nel flusso completo (che richiede anche il file acquisti)

# -------------------------------------------------------------------
# SOLO EXPORT + ACQUISTI: flat-file Amazon + file acquisti (senza inventario)
# -------------------------------------------------------------------
if export_file is not None and purchase_df is not None and inventory_df is None:
    st.subheader("Non Automated SKUs – da export Amazon (solo export + acquisti)")

    # Leggo l'export con header dinamico
    try:
        export_df = load_amazon_export(export_file)
    except Exception as e:
        st.error(f"Errore nel parsing del flat-file export: {e}")
        st.stop()

    if "SKU" not in export_df.columns:
        st.error("Il flat-file export non contiene 'SKU'.")
        st.stop()

    export_df["_SKU_KEY_"] = normalize_sku(export_df["SKU"], parse_option)

    # Individuo chiave SKU nel file acquisti
    price_key_candidates = [c for c in purchase_df.columns if c.upper() in {"CODICE","SKU","CODICE(ASIN)"}]
    if not price_key_candidates:
        st.error("Nel file acquisti manca la colonna SKU/CODICE/CODICE(ASIN).")
        st.stop()
    price_key_exp = st.selectbox("Colonna SKU nel file acquisti", price_key_candidates, index=0, key="exponly_pricekey")
    purchase_df["_SKU_KEY_"] = normalize_sku(purchase_df[price_key_exp], parse_option)

    # Trovo la colonna prezzo d'acquisto
    candidate_price_cols = [
        c for c in purchase_df.columns
        if ("prezzo" in c.lower() and "medio" in c.lower())
        or c.lower().strip() in {"prezzo","prezzo medio","prezzo medio (€)"}
    ]
    if not candidate_price_cols:
        st.error("Colonna prezzo d'acquisto non trovata (es. 'Prezzo medio').")
        st.stop()
    price_col_exp = candidate_price_cols[0] if len(candidate_price_cols) == 1 else st.selectbox("Colonna prezzo d'acquisto", candidate_price_cols, index=0, key="exponly_pricecol")

    # Categoria opzionale
    cat_cols_exp = [c for c in purchase_df.columns if c.lower().strip() == "categoria"]

    cols = ["_SKU_KEY_", price_col_exp] + (["Categoria"] if cat_cols_exp else [])
    if cat_cols_exp:
        purchase_df = purchase_df.rename(columns={cat_cols_exp[0]: "Categoria"})

    # Merge export ⟷ acquisti
    merged_export = export_df.merge(
        purchase_df[cols].rename(columns={price_col_exp: "Prezzo medio acquisto (€)"}),
        on="_SKU_KEY_", how="left", suffixes=("", "_acq")
    )

    # Parametri (indipendenti dall'inventario)
    st.markdown("### Parametri costi e margini (solo export + acquisti)")
    defaults_only = CATEGORY_MAP.get("_default", {"referral": 15.0, "closing": 0.0})
    referral_fee_pct_exp = st.number_input("% Commissione Amazon", value=defaults_only["referral"], min_value=0.0, key="exponly_ref")
    shipping_cost_exp     = st.number_input("Costo spedizione €", value=0.0, min_value=0.0, key="exponly_ship")
    vat_pct_exp           = st.number_input("IVA %", value=22.0, min_value=0.0, step=0.1, key="exponly_vat")
    margin_pct_exp        = st.number_input("Margine desiderato %", value=20.0, min_value=0.0, key="exponly_margin")
    closing_fee_exp       = defaults_only.get("closing", 0.0)

    # Calcolo minimo
    merged_export["Prezzo medio acquisto (€)"] = pd.to_numeric(merged_export["Prezzo medio acquisto (€)"], errors="coerce")
    merged_export["Prezzo minimo suggerito (€)"] = merged_export.apply(
        calc_min_price,
        axis=1,
        referral_pct=referral_fee_pct_exp,
        closing_fee=closing_fee_exp,
        dst_pct=DST_PCT,
        ship_cost=shipping_cost_exp,
        vat_pct=vat_pct_exp,
        margin_pct=margin_pct_exp,
    )

    # Opzioni flat-file
    only_missing_min_exp = st.checkbox("Mostra solo righe senza 'minimum-seller-allowed-price' nel file", value=True, key="exponly_missmin")
    overwrite_min_exp    = st.checkbox("Sovrascrivi 'minimum-seller-allowed-price' se già presente", value=False, key="exponly_overwrite")

    default_rule_base_exp = st.text_input("Rule name di default", value="AUTO", key="exponly_rulename")
    today_str_exp = pd.Timestamp.today().strftime("%Y%m%d")

    if "rule-name" not in merged_export.columns:
        merged_export["rule-name"] = ""
    if "country-code" not in merged_export.columns:
        merged_export["country-code"] = "IT"
    if "currency-code" not in merged_export.columns:
        merged_export["currency-code"] = "EUR"

    merged_export["rule-action"] = merged_export.get("rule-action", "START")
    merged_export["rule-action"] = merged_export["rule-action"].astype(str).str.upper().replace({"": "START"})

    def _mk_rule_name_exp(row):
        rn = str(row.get("rule-name") or "").strip()
        if rn:
            return rn
        cc = str(row.get("country-code") or "IT").upper()
        return f"{default_rule_base_exp}-{cc}-{today_str_exp}"

    merged_export["rule-name"] = merged_export.apply(_mk_rule_name_exp, axis=1)

    # Filtri e riempimento minimo
    if "minimum-seller-allowed-price" in merged_export.columns and only_missing_min_exp:
        view_exp = merged_export[merged_export["minimum-seller-allowed-price"].isna()].copy()
    else:
        view_exp = merged_export.copy()

    if overwrite_min_exp or "minimum-seller-allowed-price" not in view_exp.columns:
        view_exp["minimum-seller-allowed-price"] = view_exp["Prezzo minimo suggerito (€)"]
    else:
        view_exp["minimum-seller-allowed-price"] = view_exp["minimum-seller-allowed-price"].where(
            view_exp["minimum-seller-allowed-price"].notna(), view_exp["Prezzo minimo suggerito (€)"]
        )

    ff_out_exp = view_exp.dropna(subset=["SKU", "Prezzo medio acquisto (€)", "minimum-seller-allowed-price"])  # richiedo costo valido

    st.dataframe(
        ff_out_exp[[
            "SKU","country-code","currency-code","rule-name","rule-action",
            "minimum-seller-allowed-price","Prezzo medio acquisto (€)","Prezzo minimo suggerito (€)"
        ]].head(100),
        use_container_width=True, hide_index=True
    )

    st.download_button(
        "💾 Scarica Flat-File (compilato da export)",
        data=make_flatfile_bytes(build_flatfile(ff_out_exp, "SKU")),
        file_name="AutomatePricing_FromExport.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.stop()

# ---------------------------------------------------------
# Preparazione e merge
# ---------------------------------------------------------
merged_df, inv_key = get_merged_inventory(inventory_df, purchase_df, parse_option)

# Parametri costi e margini
st.subheader("Parametri costi e margini")

cats = list(CATEGORY_MAP.keys())
selected_cat = st.selectbox("Categoria", cats)
# il selettore serve solo per preimpostare le commissioni

defaults = CATEGORY_MAP.get(selected_cat, CATEGORY_MAP["_default"])
referral_fee_pct = st.number_input(
    "% Commissione Amazon", value=defaults["referral"], min_value=0.0
)
shipping_cost = st.number_input("Costo spedizione €", value=0.0, min_value=0.0)
vat_pct = st.number_input("IVA %", value=22.0, min_value=0.0, step=0.1)
margin_pct = st.number_input("Margine desiderato %", value=20.0, min_value=0.0)

show_only_matches = st.checkbox(
    "Mostra solo articoli presenti nel file acquisti", value=False
)

# verifica parametri di costo/margine
denom_check = 1 - (1 + vat_pct/100) * (
    (1 + DST_PCT/100) * (referral_fee_pct/100) + margin_pct/100
)
if denom_check <= 0:
    st.warning(
        "Parametri non validi: commissioni + margine troppo alti rispetto al prezzo."
    )

closing_fee = CATEGORY_MAP.get(selected_cat, CATEGORY_MAP["_default"])["closing"]

merged_df["Prezzo minimo suggerito (€)"] = merged_df.apply(
    calc_min_price,
    axis=1,
    referral_pct=referral_fee_pct,
    closing_fee=closing_fee,
    dst_pct=DST_PCT,
    ship_cost=shipping_cost,
    vat_pct=vat_pct,
    margin_pct=margin_pct,
)

invalid_mask = (
    merged_df["Prezzo minimo suggerito (€)"].isna()
    & merged_df["Prezzo medio acquisto (€)"].notna()
    & (merged_df["Prezzo medio acquisto (€)"] > 0)
)
if invalid_mask.any():
    st.warning(
        "Parametri non validi: commissioni + margine troppo alti rispetto al prezzo."
    )

# Applica filtro opzionale per mostrare solo gli articoli presenti nel file acquisti
display_df = merged_df
if show_only_matches:
    display_df = merged_df[merged_df["Prezzo medio acquisto (€)"].notna()]

# ---------------------------------------------------------
# Non Automated SKUs – integrazione da export Amazon
# ---------------------------------------------------------
if export_file is not None:
    st.subheader("Non Automated SKUs – da export Amazon")

    try:
        export_df = load_amazon_export(export_file)
        if "SKU" not in export_df.columns:
            st.warning("Il flat-file export non contiene la colonna 'SKU'.")
            export_df = None
    except Exception as e:
        st.error(f"Errore nel parsing del flat-file export: {e}")
        export_df = None

    if export_df is not None:
        export_df["_SKU_KEY_"] = normalize_sku(export_df["SKU"], parse_option)

        # Merge con costi dall'inventario unificato
        cols_to_pull = ["_SKU_KEY_", "Prezzo medio acquisto (€)", "Categoria"]
        cols_to_pull = [c for c in cols_to_pull if c in merged_df.columns]
        ff_df = export_df.merge(
            merged_df[cols_to_pull],
            on="_SKU_KEY_", how="left", suffixes=("", "_inv")
        )

        # Calcolo del Prezzo minimo suggerito (€) per le righe del flat-file export
        def _min_price_row(row):
            return calc_min_price(
                row=row,
                referral_pct=referral_fee_pct,
                closing_fee=closing_fee,
                dst_pct=DST_PCT,
                ship_cost=shipping_cost,
                vat_pct=vat_pct,
                margin_pct=margin_pct,
            )

        ff_df["Prezzo minimo suggerito (€)"] = ff_df.apply(_min_price_row, axis=1)

        only_missing_min = st.checkbox("Mostra solo righe senza 'minimum-seller-allowed-price' nel file", value=True)
        overwrite_min = st.checkbox("Sovrascrivi il valore esistente di 'minimum-seller-allowed-price' (se presente)", value=False)

        default_rule_base = st.text_input("Rule name di default", value="AUTO")
        today_str = pd.Timestamp.today().strftime("%Y%m%d")

        if "rule-name" not in ff_df.columns:
            ff_df["rule-name"] = ""

        if "country-code" not in ff_df.columns:
            ff_df["country-code"] = "IT"

        if "currency-code" not in ff_df.columns:
            ff_df["currency-code"] = "EUR"

        ff_df["rule-action"] = ff_df.get("rule-action", "START")
        ff_df["rule-action"] = ff_df["rule-action"].astype(str).str.upper().replace({"": "START"})

        def _mk_rule_name(row):
            rn = str(row.get("rule-name") or "").strip()
            if rn:
                return rn
            cc = str(row.get("country-code") or "IT").upper()
            return f"{default_rule_base}-{cc}-{today_str}"

        ff_df["rule-name"] = ff_df.apply(_mk_rule_name, axis=1)

        if "minimum-seller-allowed-price" in ff_df.columns and only_missing_min:
            ff_df_view = ff_df[ff_df["minimum-seller-allowed-price"].isna()].copy()
        else:
            ff_df_view = ff_df.copy()

        if overwrite_min or "minimum-seller-allowed-price" not in ff_df_view.columns:
            ff_df_view["minimum-seller-allowed-price"] = ff_df_view["Prezzo minimo suggerito (€)"]
        else:
            ff_df_view["minimum-seller-allowed-price"] = ff_df_view["minimum-seller-allowed-price"].where(
                ff_df_view["minimum-seller-allowed-price"].notna(),
                ff_df_view["Prezzo minimo suggerito (€)"]
            )

        ff_out = ff_df_view.dropna(subset=["SKU", "minimum-seller-allowed-price"])

        st.dataframe(
            ff_out[[
                "SKU",
                "country-code",
                "currency-code",
                "rule-name",
                "rule-action",
                "minimum-seller-allowed-price",
                "Prezzo medio acquisto (€)",
                "Prezzo minimo suggerito (€)",
            ]].head(100),
            use_container_width=True,
            hide_index=True,
        )

        st.download_button(
            "💾 Scarica Flat-File (compilato da export)",
            data=make_flatfile_bytes(build_flatfile(ff_out, "SKU")),
            file_name="AutomatePricing_FromExport.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# ---------------------------------------------------------
# Dashboard dei risultati
# ---------------------------------------------------------
st.subheader("Anteprima dataset unificato")

edited_df = st.data_editor(
    display_df,
    key="merged_df_editor",
    use_container_width=True,
    hide_index=True,
    column_config={
        inv_key: st.column_config.TextColumn(disabled=False)
    },
)

edited_df["_SKU_KEY_"] = normalize_sku(edited_df[inv_key], parse_option)

if st.button("🔄 Ricalcola prezzi minimi"):
    edited_df["Prezzo minimo suggerito (€)"] = edited_df.apply(
        calc_min_price,
        axis=1,
        referral_pct=referral_fee_pct,
        closing_fee=closing_fee,
        dst_pct=DST_PCT,
        ship_cost=shipping_cost,
        vat_pct=vat_pct,
        margin_pct=margin_pct,
    )

styled_df = edited_df.style.apply(highlight_below, axis=1)
st.dataframe(styled_df, use_container_width=True, hide_index=True)

# KPI sintetici
col1, col2, col3 = st.columns(3)
with col1:
    st.metric("SKU totali", len(edited_df))
with col2:
    st.metric("Prezzo medio acquisto", f"{pd.to_numeric(edited_df['Prezzo medio acquisto (€)'], errors='coerce').mean():.2f} €")
with col3:
    if "Quantita'" in edited_df.columns:
        valore = (
            pd.to_numeric(edited_df["Prezzo"], errors="coerce").fillna(0)
            * pd.to_numeric(edited_df["Quantita'"], errors="coerce").fillna(0)
        ).sum()
        st.metric("Valore inventario (listino)", f"{valore:.2f} €")
    else:
        st.metric("Valore inventario (listino)", "—")

# ---------------------------------------------------------
# Download Excel unificato + Flat-file (min price)
# ---------------------------------------------------------
output = io.BytesIO()
with pd.ExcelWriter(output, engine="openpyxl") as writer:
    excel_df = edited_df.drop(columns=["_SKU_KEY_"], errors="ignore")
    excel_df.to_excel(writer, index=False, sheet_name="inventario_match")
output.seek(0)

st.download_button(
    label="💾 Scarica Excel unificato",
    data=output,
    file_name=download_name or "inventario_match.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.download_button(
    "💾 Scarica Flat-File (min price)",
    data=make_flatfile_bytes(build_flatfile(edited_df, inv_key)),
    file_name="AutomatePricing_MinOnly.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

# ---------------------------------------------------------
# Credits
# ---------------------------------------------------------
st.caption("Made with Streamlit · Ultimo aggiornamento: 13 ago 2025")
