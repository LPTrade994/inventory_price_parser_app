import ast
import io
import types
from pathlib import Path

import pandas as pd
import numpy as np
import pytest
import openpyxl


def _sample_workbook_from_csv():
    """Return a BytesIO Excel file built from the CSV sample."""
    df = pd.read_csv(
        Path('tests/sample_modello_prezzo.csv'), sep=';', header=None, dtype=str
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Modello assegnazione prezzo'
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def _get_load_amazon_template():
    """Extract the load_amazon_template function without running Streamlit app."""
    source = Path('inventory_price_parser_app.py').read_text(encoding='utf-8')
    tree = ast.parse(source)
    nodes = []
    for node in tree.body:
        if isinstance(node, (ast.Import, ast.ImportFrom)):
            nodes.append(node)
        elif isinstance(node, ast.Assign):
            targets = [t.id for t in node.targets if isinstance(t, ast.Name)]
            if 'COUNTRY_CODES' in targets:
                nodes.append(node)
        elif isinstance(node, ast.FunctionDef) and node.name == 'load_amazon_template':
            nodes.append(node)
            break
    module = types.ModuleType('tmp')
    exec(compile(ast.Module(body=nodes, type_ignores=[]), 'inventory_price_parser_app.py', 'exec'), module.__dict__)
    return module.load_amazon_template


@pytest.fixture(scope='module')
def load_amazon_template():
    return _get_load_amazon_template()


def test_load_amazon_template_basic(load_amazon_template):
    df = load_amazon_template(_sample_workbook_from_csv())
    assert list(df.columns) == [
        'SKU',
        'Prezzo',
        'rule-action',
        'country-code',
        'sales-rank',
        'customer-views-share',
    ]
    assert len(df) == 2
    assert df['sales-rank'].tolist() == [100, pd.NA]
    assert df['customer-views-share'].iloc[0] == pytest.approx(0.5)
    assert np.isnan(df['customer-views-share'].iloc[1])


def test_load_amazon_template_invalid_rule_action(load_amazon_template):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Modello assegnazione prezzo'
    ws.append(['H'] * 4)
    ws.append(['SKU', 'Current Selling Price', 'Rule action', 'Country Code'])
    ws.append(['X1', '1', 'INVALID', 'IT'])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with pytest.raises(ValueError, match="rule-action"):
        load_amazon_template(buf)


def test_load_amazon_template_invalid_country_code(load_amazon_template):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Modello assegnazione prezzo'
    ws.append(['H'] * 4)
    ws.append(['SKU', 'Current Selling Price', 'Rule action', 'Country Code'])
    ws.append(['X1', '1', 'START', 'XX'])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with pytest.raises(ValueError, match="Codici paese"):
        load_amazon_template(buf)
